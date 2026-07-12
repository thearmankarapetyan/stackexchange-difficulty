"""This module builds the canonical project data-dictionary workbook.

The script combines the complete characteristic catalogue, the implemented
question-table schema, and the tracked pilot table. It creates one XLSX file
with an overview, the complete catalogue, the current output contract, and a
small current sample.

Example:
    python src/build_data_dictionary.py

Alternative input and output paths are available through command-line
arguments. The script requires Python 3.10 or newer and ``openpyxl`` from
``requirements-dev.txt``. Input TSV files remain unchanged, and the workbook
is replaced atomically after successful construction.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_CATALOGUE = PROJECT_DIR / "config" / "characteristic_catalogue.tsv"
DEFAULT_SCHEMA = PROJECT_DIR / "config" / "characteristics.tsv"
DEFAULT_SAMPLE = PROJECT_DIR / "data" / "examples" / "characteristics_pilot.tsv"
DEFAULT_OUTPUT = PROJECT_DIR / "docs" / "reference" / "data-dictionary.xlsx"
DICTIONARY_V5_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1flXjpBsRzzyYj2n0BAqaOWjoDZgNNVnKv6tDsGSJfBI"
)
DICTIONARY_V5_RETRIEVED = "12 July 2026"
DICTIONARY_V5_SHA256 = (
    "48b243fc57c9d36c577f6daf49867b3a02d24cd9a74a299bc4fe462459cfc218"
)

CATALOGUE_COLUMNS = [
    "catalogue_id",
    "characteristic",
    "source_name_v5",
    "entity",
    "implementation_status",
    "current_output_field",
    "characteristic_group",
    "availability_stage",
    "role",
    "source",
    "data_type",
    "unit_or_values",
    "definition",
    "calculation",
    "interpretation",
    "empty_value_meaning",
    "evidence_or_reference",
    "notes",
]
CATALOGUE_LABELS = [
    "Catalogue ID",
    "Canonical characteristic",
    "Dictionary v5 source name",
    "Entity",
    "Implementation status",
    "Current output field",
    "Calculation group",
    "Availability stage",
    "Analytical role",
    "Exact source or required evidence",
    "Data type",
    "Unit or allowed values",
    "Definition",
    "Calculation method",
    "Interpretation",
    "Empty-value meaning",
    "Scientific evidence or reference",
    "Synthesis notes",
]
IMPLEMENTATION_STATUSES = {
    "Implemented",
    "Available in current source",
    "Requires additional dump file",
    "Requires Data Explorer source",
    "Requires manual assessment",
    "Requires model evaluation",
    "Literature-derived candidate",
    "Needs source review",
}
CALCULATION_GROUPS = {
    "Non-calculated",
    "Calculated by Stack Exchange",
    "Calculated by project",
    "Assessed manually",
    "Calculated during model evaluation",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_GRAY = Side(style="thin", color="D9E1F2")


def read_tsv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """Returns the header and rows from one UTF-8 tab-separated file."""
    if not path.is_file():
        raise FileNotFoundError(f"TSV file not found: {path}")
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        if reader.fieldnames is None:
            raise ValueError(f"TSV file has no header: {path}")
        return reader.fieldnames, list(reader)


def validate_inputs(
    catalogue_header: list[str],
    catalogue: list[dict[str, str]],
    schema_header: list[str],
    schema: list[dict[str, str]],
    sample_header: list[str],
) -> None:
    """Validates catalogue uniqueness and current-output synchronization."""
    if catalogue_header != CATALOGUE_COLUMNS:
        raise ValueError("Characteristic catalogue columns do not match the contract")
    if len(CATALOGUE_LABELS) != len(CATALOGUE_COLUMNS):
        raise ValueError("Characteristic catalogue workbook labels are incomplete")
    if not catalogue:
        raise ValueError("Characteristic catalogue must contain at least one row")
    required_schema_columns = {
        "position", "characteristic", "characteristic_group",
        "availability_stage", "role", "source", "data_type",
        "unit_or_values", "definition", "calculation", "interpretation",
        "empty_value_meaning",
    }
    if not required_schema_columns.issubset(schema_header):
        raise ValueError("Current schema columns do not match the contract")

    optional_catalogue_values = {
        "source_name_v5", "current_output_field", "evidence_or_reference", "notes"
    }
    for row in catalogue:
        empty = [
            column
            for column in CATALOGUE_COLUMNS
            if column not in optional_catalogue_values and not row[column].strip()
        ]
        if empty:
            raise ValueError(
                f"Catalogue row {row['catalogue_id'] or '<empty>'} has empty "
                + ", ".join(empty)
            )

    catalogue_ids = [row["catalogue_id"] for row in catalogue]
    names = [row["characteristic"] for row in catalogue]
    expected_ids = [f"C{number:03d}" for number in range(1, len(catalogue) + 1)]
    if catalogue_ids != expected_ids:
        raise ValueError("Catalogue IDs must be consecutive and ordered from C001")
    if len(catalogue_ids) != len(set(catalogue_ids)):
        raise ValueError("Characteristic catalogue IDs must be unique")
    if len(names) != len(set(names)):
        raise ValueError("Characteristic names must be unique")
    statuses = {row["implementation_status"] for row in catalogue}
    if not statuses.issubset(IMPLEMENTATION_STATUSES):
        raise ValueError("Characteristic catalogue contains an unknown status")
    groups = {row["characteristic_group"] for row in catalogue}
    if not groups.issubset(CALCULATION_GROUPS):
        raise ValueError("Characteristic catalogue contains an unknown calculation group")
    v5_names = [
        name
        for row in catalogue
        for name in row["source_name_v5"].split("; ")
        if name
    ]
    if len(v5_names) != len(set(v5_names)):
        raise ValueError("Dictionary v5 source names must map to one catalogue row")

    implemented_fields = [
        row["current_output_field"]
        for row in catalogue
        if row["implementation_status"] == "Implemented"
    ]
    if any(not field for field in implemented_fields):
        raise ValueError("Every implemented catalogue row requires an output field")
    if len(implemented_fields) != len(set(implemented_fields)):
        raise ValueError("Implemented catalogue output fields must be unique")
    if any(
        row["current_output_field"]
        for row in catalogue
        if row["implementation_status"] != "Implemented"
    ):
        raise ValueError("Only implemented catalogue rows can map to current output fields")
    implemented = set(implemented_fields)
    schema_names = [row["characteristic"] for row in schema]
    if list(range(1, len(schema) + 1)) != [int(row["position"]) for row in schema]:
        raise ValueError("Current schema positions must be consecutive from 1")
    if len(schema_names) != len(set(schema_names)):
        raise ValueError("Current schema characteristic names must be unique")
    if set(schema_names) != implemented:
        raise ValueError("The current schema and implemented catalogue subset differ")
    if sample_header != schema_names:
        raise ValueError("Pilot sample columns do not match the current schema order")


def style_table_sheet(
    sheet: Worksheet,
    rows: list[list[str]],
    table_name: str,
    *,
    wrap_text: bool = True,
    maximum_width: int = 42,
    pages_wide: int = 2,
) -> None:
    """Adds rows, readable formatting, filters, and bounded column widths."""
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=wrap_text, vertical="top")
            cell.border = Border(bottom=THIN_GRAY)
        if not wrap_text:
            sheet.row_dimensions[row[0].row].height = 30

    for column_number, column in enumerate(sheet.columns, start=1):
        longest = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[get_column_letter(column_number)].width = min(
            max(longest + 2, 12), maximum_width
        )

    reference = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    table = Table(displayName=table_name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = pages_wide
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:1"
    sheet.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.4, bottom=0.4, header=0.15, footer=0.15
    )


def add_overview(
    workbook: Workbook,
    catalogue: list[dict[str, str]],
    schema: list[dict[str, str]],
    sample: list[dict[str, str]],
) -> None:
    """Creates the workbook overview and status guide."""
    v5_source_names = [
        name
        for row in catalogue
        for name in row["source_name_v5"].split("; ")
        if name
    ]
    v5_concepts = sum(bool(row["source_name_v5"]) for row in catalogue)
    production_only_concepts = len(catalogue) - v5_concepts
    sheet = workbook.active
    sheet.title = "Overview"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 31
    sheet.column_dimensions["B"].width = 105

    rows = [
        ("Project data dictionary", "Complete characteristic catalogue and current implemented output contract"),
        ("Catalogue scope", f"{len(catalogue)} distinct characteristics synthesized from Dictionary v5 and the verified project schema."),
        (
            "Dictionary v5 accounting",
            f"{len(v5_source_names)} source names map to {v5_concepts} distinct "
            "concepts. fastest_response_time_hours and "
            "time_to_first_answer_hours describe the same earliest-answer delay "
            f"and share one catalogue row. {production_only_concepts} verified "
            "production concepts absent from Dictionary v5 complete the catalogue.",
        ),
        ("Current output", f"{len(schema)} implemented question-level fields produced by build_characteristics.py."),
        (
            "Dictionary v5 output additions",
            "code_character_count and has_stackexchange_answer are direct, "
            "meaningful question-level values. Other currently available v5 "
            "source fields describe repeated answer, comment, vote, or tag "
            "records and require a separate table or an explicit aggregation.",
        ),
        ("Source catalogue", f"Dictionary v5: {DICTIONARY_V5_URL}"),
        ("Source retrieval date", DICTIONARY_V5_RETRIEVED),
        ("Source export SHA-256", DICTIONARY_V5_SHA256),
        ("Scientific references", "Evidence labels refer to docs/explanation/state-of-the-art-qpp-ppp-rag.pdf."),
        ("Sheet: Characteristic catalogue", "Every retained concept, its entity, status, source, definition, calculation, interpretation, empty-value meaning, evidence, and synthesis note."),
        ("Sheet: Current output", f"The {len(schema)} implemented fields in generated TSV order, with their catalogue mapping and calculation contract."),
        ("Sheet: Current sample", f"The tracked {len(sample)}-question pilot generated by the current implementation."),
        ("Navigation", "Table filters support selection by entity, implementation status, calculation group, source, or evidence. Frozen headers remain visible during scrolling."),
        ("Output inclusion", "An implemented field has one unambiguous value per selected question, an available source, and a complete calculation contract."),
        ("Status: Implemented", "The current builder produces the field and verifies it against the current schema."),
        ("Status: Available in current source", "The required information exists in Posts.xml, Comments.xml, or Votes.xml and remains outside the current question-level output."),
        (
            "Status: Requires additional dump file",
            "The characteristic requires another official public-dump XML file "
            "beyond the current Posts.xml, Comments.xml, and Votes.xml inputs.",
        ),
        (
            "Status: Requires Data Explorer source",
            "The required Stack Exchange table is available through Stack "
            "Exchange Data Explorer and absent from the official public XML dump.",
        ),
        ("Status: Requires manual assessment", "The characteristic requires a documented human-review protocol and retained assessment evidence."),
        ("Status: Requires model evaluation", "The characteristic requires model outputs and a documented evaluation protocol."),
        ("Status: Literature-derived candidate", "The characteristic comes from the reviewed literature and requires scientific selection plus a complete implementation specification."),
        ("Status: Needs source review", "The source field requires availability and meaning verification for the selected dump."),
        ("Calculation groups", "Non-calculated; Calculated by Stack Exchange; Calculated by project; Assessed manually; Calculated during model evaluation."),
        ("Empty values", "Each catalogue row states the exact meaning of an empty value. Zero and FALSE remain observed values when the row definition identifies them as valid."),
        (
            "Synchronization",
            "config/characteristic_catalogue.tsv is the canonical full catalogue. "
            f"config/characteristics.tsv is the canonical {len(schema)}-field "
            "production schema.",
        ),
    ]
    for row_number, (label, value) in enumerate(rows, start=1):
        sheet.cell(row=row_number, column=1, value=label)
        sheet.cell(row=row_number, column=2, value=value)
        sheet.cell(row=row_number, column=1).font = Font(bold=True)
        sheet.cell(row=row_number, column=1).fill = SECTION_FILL
        for column in (1, 2):
            sheet.cell(row=row_number, column=column).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
            sheet.cell(row=row_number, column=column).border = Border(bottom=THIN_GRAY)
    sheet.row_dimensions[1].height = 31
    sheet["A1"].font = Font(bold=True, size=15, color="1F4E78")
    sheet["B1"].font = Font(bold=True, size=13, color="1F4E78")
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins = PageMargins(
        left=0.35, right=0.35, top=0.45, bottom=0.45, header=0.15, footer=0.15
    )


def build_workbook(
    catalogue: list[dict[str, str]],
    schema: list[dict[str, str]],
    sample_header: list[str],
    sample: list[dict[str, str]],
) -> Workbook:
    """Builds the four-sheet workbook in memory."""
    workbook = Workbook()
    workbook.properties.title = "Stack Exchange characteristic data dictionary"
    workbook.properties.subject = "Complete catalogue and current output contract"
    workbook.properties.creator = "Stack Exchange difficulty project"
    workbook.properties.created = datetime(2026, 7, 12)
    workbook.properties.modified = datetime(2026, 7, 12)
    add_overview(workbook, catalogue, schema, sample)

    catalogue_sheet = workbook.create_sheet("Characteristic catalogue")
    style_table_sheet(
        catalogue_sheet,
        [CATALOGUE_LABELS]
        + [[row[column] for column in CATALOGUE_COLUMNS] for row in catalogue],
        "CharacteristicCatalogue",
    )

    by_name = {row["characteristic"]: row for row in catalogue}
    output_labels = [
        "Position", "Current output field", "Catalogue ID", "Entity",
        "Calculation group", "Availability stage", "Analytical role",
        "Exact source", "Data type", "Unit or allowed values", "Definition",
        "Calculation method", "Interpretation", "Empty-value meaning",
        "Implementation file",
    ]
    output_rows = []
    for schema_row in schema:
        entry = by_name[schema_row["characteristic"]]
        output_rows.append([
            schema_row["position"], schema_row["characteristic"],
            entry["catalogue_id"], entry["entity"],
            schema_row["characteristic_group"], schema_row["availability_stage"],
            schema_row["role"], schema_row["source"], schema_row["data_type"],
            schema_row["unit_or_values"], schema_row["definition"],
            schema_row["calculation"], schema_row["interpretation"],
            schema_row["empty_value_meaning"],
            "src/question_characteristics.py",
        ])
    current_sheet = workbook.create_sheet("Current output")
    style_table_sheet(
        current_sheet,
        [output_labels] + output_rows,
        "CurrentOutput",
    )

    sample_sheet = workbook.create_sheet("Current sample")
    style_table_sheet(
        sample_sheet,
        [sample_header]
        + [[row[column] for column in sample_header] for row in sample],
        "CurrentSample",
        wrap_text=False,
        maximum_width=24,
        pages_wide=3,
    )
    return workbook


def workbook_signature(workbook: Workbook) -> tuple[object, ...]:
    """Returns workbook values and maintained layout settings for comparison."""
    sheets = []
    for sheet in workbook.worksheets:
        values = tuple(
            tuple(cell.value if cell.value is not None else "" for cell in row)
            for row in sheet.iter_rows(
                min_row=1,
                max_row=sheet.max_row,
                min_col=1,
                max_col=sheet.max_column,
            )
        )
        widths = tuple(
            (letter, sheet.column_dimensions[letter].width)
            for letter in sorted(sheet.column_dimensions)
        )
        tables = tuple(
            (table.name, table.ref, table.tableStyleInfo.name)
            for table in sheet.tables.values()
        )
        sheets.append(
            (
                sheet.title,
                values,
                str(sheet.freeze_panes or ""),
                sheet.sheet_view.showGridLines,
                widths,
                tables,
                sheet.page_setup.orientation,
                sheet.page_setup.paperSize,
                sheet.page_setup.fitToWidth,
                sheet.page_setup.fitToHeight,
                sheet.print_title_rows,
            )
        )
    return tuple(sheets)


def parse_args(arguments: Sequence[str] | None = None) -> argparse.Namespace:
    """Returns configurable source and destination paths."""
    parser = argparse.ArgumentParser(
        description="Canonical project data-dictionary workbook construction."
    )
    parser.add_argument("--catalogue", type=Path, default=DEFAULT_CATALOGUE)
    parser.add_argument("--schema", type=Path, default=DEFAULT_SCHEMA)
    parser.add_argument("--sample", type=Path, default=DEFAULT_SAMPLE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--check",
        action="store_true",
        help="Verification that the destination workbook matches the current TSV inputs.",
    )
    return parser.parse_args(arguments)


def run(arguments: Sequence[str] | None = None) -> Path:
    """Validates the TSV inputs and publishes the generated workbook."""
    args = parse_args(arguments)
    catalogue_header, catalogue = read_tsv(args.catalogue)
    schema_header, schema = read_tsv(args.schema)
    sample_header, sample = read_tsv(args.sample)
    validate_inputs(
        catalogue_header, catalogue, schema_header, schema, sample_header
    )
    workbook = build_workbook(catalogue, schema, sample_header, sample)

    if args.check:
        if not args.output.is_file():
            raise FileNotFoundError(f"Workbook not found: {args.output}")
        existing = load_workbook(args.output, data_only=False)
        if workbook_signature(existing) != workbook_signature(workbook):
            raise ValueError(
                f"Workbook is out of date with the canonical TSV inputs: {args.output}"
            )
        print(f"Verified {args.output}")
        return args.output

    args.output.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", delete=False, dir=args.output.parent
        ) as temporary:
            temporary_path = Path(temporary.name)
        workbook.save(temporary_path)
        os.replace(temporary_path, args.output)
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
    print(f"Wrote {args.output}")
    return args.output


def main(arguments: Sequence[str] | None = None) -> int:
    """Runs the command-line program and reports a contextual error."""
    try:
        run(arguments)
        return 0
    except (FileNotFoundError, OSError, ValueError) as error:
        print(f"error: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
